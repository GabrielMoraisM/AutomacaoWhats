import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from urllib.parse import quote
from time import sleep

def Automatizacao(mensagens, caminho_arquivo, nome_planilha, nome_pagina, caminho_foto):
    navegador = None
    try:
        # Inicialize o navegador
        navegador = webdriver.Chrome()
        navegador.get("https://web.whatsapp.com/")

        # Aguarde até que o WhatsApp Web carregue
        WebDriverWait(navegador, 120).until(EC.presence_of_element_located((By.ID, "side")))
        print("WhatsApp Web carregado com sucesso.")

        # Carregar a planilha e a página específica
        workbook = openpyxl.load_workbook(caminho_arquivo)
        paginaContatos = workbook[nome_pagina]

        for linha in paginaContatos.iter_rows(min_row=2):
            if len(linha) < 3:
                print(f"Linha inválida ou incompleta encontrada: {linha}")
                continue  # Pule para a próxima linha

            nome = linha[0].value
            telefone = str(linha[1].value).strip()  # Converta para string e aplique strip()
            nomeReal = linha[2].value
            print(f"Mensagem aberta na conta de : {nomeReal}")

            if not telefone:
                print(f"Telefone inválido para o contato {nome}. Pulando envio.")
                continue

            # Envio das mensagens
            for mensagem in mensagens:
                mensagem_personalizada = mensagem.replace("{nome_lead}", nome)
                linkWhats = f"https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem_personalizada)}"

                navegador.get(linkWhats)
                sleep(20)

                try:
                    campo_mensagem = WebDriverWait(navegador, 45).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[1]/div/div/p[1]/span'))
                    )
                    campo_mensagem.send_keys(Keys.ENTER)
                    sleep(5)
                    
                    print(f"Mensagem enviada para {nomeReal} ({telefone})")
                    
                except Exception as e:
                    print(f"Erro ao enviar mensagem para {nomeReal} ({telefone}): {e}")

            # Envio da imagem (fora do loop de mensagens)
            if caminho_foto:
                try:
                    campo_anexar = WebDriverWait(navegador, 45).until(
                        EC.presence_of_element_located((By.XPATH, '//div[@title="Anexar"]'))
                    )
                    campo_anexar.click()

                    foto_input = WebDriverWait(navegador, 45).until(
                        EC.presence_of_element_located((By.XPATH, '//input[@accept="image/*,video/mp4,video/3gpp,video/quicktime"]'))
                    )
                    foto_input.send_keys(caminho_foto)

                    campo_foto = WebDriverWait(navegador, 45).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="app"]/div/div[2]/div[2]/div[2]/span/div/div/div/div[2]/div/div[1]/div[3]/div/div/div[2]/div[1]/div[1]/p'))
                    )
                    campo_foto.send_keys(Keys.ENTER)

                    print(f"Imagem enviada para {nome} ({telefone})")

                except Exception as e:
                    print(f"Erro ao enviar imagem para {nome} ({telefone}): {e}")

            sleep(20)

    except Exception as e:
        print(f"Erro na automação: {e}")

    finally:
        if navegador:
            navegador.quit()

def reenvio(mensagens, caminho_arquivo, nome_planilha, nome_pagina, caminho_foto):
    navegador = None
    try:
        # Inicialize o navegador
        navegador = webdriver.Chrome()
        navegador.get("https://web.whatsapp.com/")

        # Aguarde até que o WhatsApp Web carregue
        WebDriverWait(navegador, 60).until(EC.presence_of_element_located((By.ID, "side")))
        print("WhatsApp Web carregado com sucesso.")

        # Carregar a planilha e a página específica
        workbook = openpyxl.load_workbook(caminho_arquivo)
        paginaContatos = workbook[nome_pagina]

        for linha in paginaContatos.iter_rows(min_row=2):
            if len(linha) < 3:
                print(f"Linha inválida ou incompleta encontrada: {linha}")
                continue  # Pule para a próxima linha

            nome = linha[0].value
            telefone = str(linha[1].value).strip()  # Converta para string e aplique strip()
            print(f"Nome contato: {nome}")

            if not telefone:
                print(f"Telefone inválido para o contato {nome}. Pulando envio.")
                continue

            if status == "pendente":
                for mensagem in mensagens:
                    mensagem_personalizada = mensagem.replace("{nome_lead}", nome).replace("{assessor}", assessor)
                    linkWhats = f"https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem_personalizada)}"
    
                    navegador.get(linkWhats)
                    sleep(20)

                    try:
                        campo_mensagem = WebDriverWait(navegador, 45).until(
                            EC.presence_of_element_located((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[1]/div/div/p[1]/span'))
                        )
                        campo_mensagem.send_keys(Keys.ENTER)
                        sleep(5)
                        
                        print(f"Mensagem enviada para {nome} ({telefone})")
                        
                    except Exception as e:
                        print(f"Erro ao enviar mensagem para {nome} ({telefone}): {e}")

                    # Envio da imagem (fora do loop de mensagens)
                if caminho_foto:
                        try:
                            campo_anexar = WebDriverWait(navegador, 45).until(
                                EC.presence_of_element_located((By.XPATH, '//div[@title="Anexar"]'))
                            )
                            campo_anexar.click()

                            foto_input = WebDriverWait(navegador, 45).until(
                                EC.presence_of_element_located((By.XPATH, '//input[@accept="image/*,video/mp4,video/3gpp,video/quicktime"]'))
                            )
                            foto_input.send_keys(caminho_foto)

                            campo_foto = WebDriverWait(navegador, 45).until(
                                EC.presence_of_element_located((By.XPATH, '//*[@id="app"]/div/div[2]/div[2]/div[2]/span/div/div/div/div[2]/div/div[1]/div[3]/div/div/div[2]/div[1]/div[1]/p'))
                            )
                            campo_foto.send_keys(Keys.ENTER)

                            print(f"Imagem enviada para {nome} ({telefone})")    

                        except Exception as e:
                            print(f"Erro ao enviar imagem para {nome} ({telefone}): {e}")

                        sleep(20)
                else:
                    continue
    except Exception as e:
        print(f"Erro na automação: {e}")

    finally:
        if navegador:
            navegador.quit()